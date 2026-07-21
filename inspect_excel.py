import os
import openpyxl
from pathlib import Path
p = Path(r'\\wsl.localhost\Ubuntu\home\ntomo\portfolio\cospace-link\個人開発_コワーキング予約システム_要件シート.xlsx')
print('resolved', p)
print('exists', p.exists())
wb = openpyxl.load_workbook(p, data_only=True)
print('sheets', wb.sheetnames)
for ws in wb.worksheets:
    print('\n===', ws.title, '===')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True):
        print(row)
